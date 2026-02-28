import os
import io
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
import openpyxl
from openpyxl.drawing.image import Image as xlImage
from openpyxl.styles import Font, Alignment, PatternFill
from models import db, Usuario, Secretaria, Contratacao, Ente
from dotenv import load_dotenv
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadTimeSignature

load_dotenv() # Carrega as variáveis do arquivo .env

app = Flask(__name__)

app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'chave-padrao-desenvolvimento')

# Cria a pasta estática de uploads se não existir
UPLOAD_FOLDER = os.path.join('static', 'img')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Configuração de Acesso ao Banco de Dados (MySQL 9.1)
DB_USER = os.environ.get('DB_USER', 'root')
DB_PASS = os.environ.get('DB_PASSWORD', '')
DB_HOST = os.environ.get('DB_HOST', 'localhost')
DB_PORT = os.environ.get('DB_PORT', '3307')
DB_NAME = os.environ.get('DB_NAME', 'pca_sdlc')

# A MÁGICA: Forçamos a verificação da variável de ambiente
if os.environ.get('AMBIENTE_DE_TESTE') == 'True' or os.environ.get('CI') == 'true':
    # Usa banco em memória (SQLite) para testes locais ou no GitHub Actions
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///:memory:'
else:
    # Usa MySQL para desenvolvimento local ou Produção
    app.config['SQLALCHEMY_DATABASE_URI'] = f"mysql+pymysql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=5)

# ============================================================================
# CONFIGURAÇÃO DE E-MAIL (Flask-Mail)
# ============================================================================
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'True') == 'True'
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_USERNAME')

mail = Mail(app)
# Gerador de tokens seguros usando a chave mestra da aplicação
s = URLSafeTimedSerializer(app.secret_key)

# Vincula o banco de dados à aplicação Flask
db.init_app(app)

# Injeta os dados do Ente e a data da última contratação em TODOS os arquivos HTML
# Injeta os dados do Ente, data e o ano atual em TODOS os arquivos HTML
# Injeta os dados Globais em TODOS os arquivos HTML
@app.context_processor
def injetar_dados_globais():
    ente = Ente.query.first()
    
    # 1. HORA DA ÚLTIMA ATUALIZAÇÃO DO BANCO DE DADOS (Usado no site)
    ultima_modificacao = Contratacao.query.order_by(Contratacao.data_atualizacao.desc()).first()
    data_ultima_modificacao = ultima_modificacao.data_atualizacao if ultima_modificacao else None
    
    # 2. HORA DE AGORA DO SERVIDOR (Usado no PDF e no ano atual)
    data_hora_agora = datetime.now()
    ano_atual = data_hora_agora.year
    
    return dict(
        ente=ente, 
        data_ultima_modificacao=data_ultima_modificacao, 
        ano_atual=ano_atual, 
        data_hora_agora=data_hora_agora
    )

# ============================================================================
# FILTROS CUSTOMIZADOS (LOCALIZAÇÃO PT-BR)
# ============================================================================
@app.template_filter('moeda_br')
def formatar_moeda(valor):
    if valor is None:
        return "0,00"
    # Coloca vírgula nos milhares e ponto no decimal, depois inverte para o padrão BR
    return "{:,.2f}".format(valor).replace(",", "X").replace(".", ",").replace("X", ".")

# ============================================================================
# PROTEÇÃO CONTRA CACHE DE NAVEGADOR (SEGURANÇA)
# ============================================================================
@app.after_request
def add_header(response):
    """
    Força o navegador a não guardar cache de nenhuma página.
    Impede que o usuário use o botão 'Voltar' do navegador após o logout.
    """
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

# ============================================================================
# INTERFACE PÚBLICA (TRANSPARÊNCIA)
# ============================================================================

@app.route('/')
def home():
    secretarias = Secretaria.query.all()
    
    # Filtros recebidos da URL
    sec_id = request.args.get('secretaria')
    exercicio = request.args.get('exercicio')
    codigo = request.args.get('codigo')

    query = Contratacao.query
    if sec_id: query = query.filter_by(secretaria_id=sec_id)
    if exercicio: query = query.filter_by(exercicio=exercicio)
    if codigo: query = query.filter(Contratacao.codigo_identificador.like(f"%{codigo}%"))

    contratacoes = query.all()
    return render_template('home.html', contratacoes=contratacoes, secretarias=secretarias)

# ============================================================================
# EXPORTAÇÃO DE RELATÓRIOS (EXCEL E PDF) OFICIAIS
# ============================================================================

def obter_dados_filtrados():
    """Lê os filtros da URL e busca os dados exatos que o cidadão está vendo na tela."""
    sec_id = request.args.get('secretaria')
    exercicio = request.args.get('exercicio')
    codigo = request.args.get('codigo')
    
    query = Contratacao.query
    if sec_id: query = query.filter_by(secretaria_id=sec_id)
    if exercicio: query = query.filter_by(exercicio=exercicio)
    if codigo: query = query.filter(Contratacao.codigo_identificador.like(f"%{codigo}%"))
        
    orgao_nome = "Consolidado (Todas as Secretarias)"
    if sec_id and sec_id != 'Todas':
        sec = Secretaria.query.get(int(sec_id))
        if sec: orgao_nome = sec.nome
            
    contratacoes = query.all()
    ente = Ente.query.first()
    return contratacoes, ente, exercicio, orgao_nome

@app.route('/exportar/excel')
def exportar_excel():
    contratacoes, ente, exercicio, orgao_nome = obter_dados_filtrados()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PCA_Exportacao"

    # 1. CABEÇALHO DO ÓRGÃO
    ws.merge_cells('B1:F1')
    ws['B1'] = ente.nome if ente else 'Órgão Público'
    ws['B1'].font = Font(size=14, bold=True, color="24549C")
    
    ws.merge_cells('B2:F2')
    ws['B2'] = f"Plano de Contratações Anual - {exercicio or 'Geral'}"
    ws['B2'].font = Font(size=12, bold=True)

    ws.merge_cells('B3:F3')
    ws['B3'] = f"Escopo: {orgao_nome}"
    
    # Altura das linhas do cabeçalho
    for i in range(1, 5): ws.row_dimensions[i].height = 20

    # 2. INSERIR LOGOTIPO (Se existir)
    if ente and ente.logo_path:
        logo_full_path = os.path.join(app.root_path, 'static', ente.logo_path)
        if os.path.exists(logo_full_path):
            try:
                img = xlImage(logo_full_path)
                img.width, img.height = 75, 75
                ws.add_image(img, 'A1')
            except Exception as e:
                print(f"Aviso: Não foi possível anexar a imagem no Excel: {e}")

    # 3. CABEÇALHO DA TABELA
    headers = ['Código', 'Exercício', 'Secretaria', 'Objeto', 'Data Planejada', 'Dotação', 'Valor Estimado (R$)']
    row_num = 5
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header_title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="24549C", end_color="24549C", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    # Alargando a coluna do Objeto
    ws.column_dimensions['D'].width = 60

    # 4. PREENCHER DADOS
    # Criamos o estilo de alinhamento centralizado uma única vez para otimizar a memória
    center_alignment = Alignment(horizontal="center", vertical="center")

    for c in contratacoes:
        row_num += 1
        
        # Coluna 1: Código (Centralizado)
        cell_codigo = ws.cell(row=row_num, column=1, value=c.codigo_identificador)
        cell_codigo.alignment = center_alignment
        
        # Coluna 2: Exercício (Centralizado)
        cell_exercicio = ws.cell(row=row_num, column=2, value=c.exercicio)
        cell_exercicio.alignment = center_alignment
        
        # Coluna 3: Secretaria (Padrão/Esquerda)
        ws.cell(row=row_num, column=3, value=c.secretaria.nome)
        
        # Coluna 4: Objeto (Padrão/Esquerda)
        ws.cell(row=row_num, column=4, value=c.objeto)
        
        # Coluna 5: Data Planejada (Centralizado)
        cell_data = ws.cell(row=row_num, column=5, value=c.data_planejada.strftime('%d/%m/%Y') if c.data_planejada else '-')
        cell_data.alignment = center_alignment
        
        # Coluna 6: Dotação (Centralizado)
        cell_dotacao = ws.cell(row=row_num, column=6, value=c.dotacao)
        cell_dotacao.alignment = center_alignment
        
        # Coluna 7: Valor Estimado (Formato Contábil, Excel alinha à direita nativamente)
        cell_valor = ws.cell(row=row_num, column=7, value=float(c.valor_estimado))
        cell_valor.number_format = 'R$ #,##0.00'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'PCA_{exercicio or "Completo"}.xlsx'
    )

@app.route('/exportar/pdf')
def exportar_pdf():
    contratacoes, ente, exercicio, orgao_nome = obter_dados_filtrados()
    return render_template('relatorio_pdf.html', contratacoes=contratacoes, ente=ente, exercicio=exercicio, orgao_nome=orgao_nome)

# ============================================================================
# INTERFACE ADMINISTRATIVA
# ============================================================================

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    # 1. Busca os dados da prefeitura/ente no banco de dados primeiro
    ente_atual = Ente.query.first()
    if request.method == 'POST':
        # TOLERÂNCIA A FALHAS: Tenta pegar pelo nome em português. Se não achar, pega pelo padrão inglês.
        login_form = request.form.get('login') or request.form.get('username')
        senha_form = request.form.get('senha') or request.form.get('password')
        
        user = Usuario.query.filter_by(login=login_form).first()
        
        senha_valida = False
        if user and senha_form:
            try:
                # Tenta usar o método do seu models.py
                senha_valida = user.check_password(senha_form)
            except AttributeError:
                # Se o método não existir no seu modelo, usa a validação padrão do framework
                from werkzeug.security import check_password_hash
                senha_valida = check_password_hash(user.senha, senha_form)

        # Se o usuário existir e a senha for válida...
        if user and senha_valida:
            session.clear()          # Prevenção contra Session Fixation (Sessão Zumbi)
            session.permanent = True # Ativa o Timeout de 5 minutos
            session['user_id'] = user.id
            session['user_login'] = user.login
            session['secretaria_id'] = user.secretaria_id
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Login ou senha incorretos. Verifique suas credenciais.')
            
    return render_template('admin_login.html', ente=ente_atual)

@app.route('/admin/logout')
def admin_logout():
    session.clear() # Limpa absolutamente tudo da sessão no backend
    return redirect(url_for('home'))

@app.route('/admin/esqueci-senha', methods=['POST'])
def esqueci_senha():
    email = request.form.get('email')
    usuario = Usuario.query.filter_by(email=email).first()
    
    if usuario:
        # Gera um token com o e-mail do usuário embutido
        token = s.dumps(usuario.email, salt='recuperacao-senha')
        # Cria o link completo apontando para a nossa rota de reset
        link = url_for('resetar_senha_token', token=token, _external=True)
        
        msg = Message('Recuperação de Senha - PCA', recipients=[email])
        msg.body = f'''Olá {usuario.nome},

Você solicitou a recuperação da sua senha no sistema PCA.
Para redefinir sua credencial, clique no link abaixo:

{link}

Este link expira em 30 minutos.
Se você não solicitou esta alteração, apenas ignore este e-mail.
'''
        try:
            mail.send(msg)
            flash('Se o e-mail estiver cadastrado, um link de recuperação será enviado. Verifique sua caixa de entrada e spam.')
        except Exception as e:
            flash('Erro interno ao tentar enviar o e-mail. Avise o suporte técnico.')
            print(f"Erro SMTP: {e}")
    else:
        # Regra de Segurança: Nunca revele se o e-mail existe ou não na base de dados
        flash('Se o e-mail estiver cadastrado, um link de recuperação será enviado. Verifique sua caixa de entrada e spam.')
        
    return redirect(url_for('admin_login'))

@app.route('/admin/resetar/<token>', methods=['GET', 'POST'])
def resetar_senha_token(token):
    try:
        # Tenta extrair o e-mail do token e verifica se passou de 30 min (1800 segundos)
        email = s.loads(token, salt='recuperacao-senha', max_age=1800)
    except SignatureExpired:
        flash('Erro: O link de recuperação expirou. Solicite um novo na tela de login.')
        return redirect(url_for('admin_login'))
    except BadTimeSignature:
        flash('Erro: Link de recuperação inválido ou corrompido.')
        return redirect(url_for('admin_login'))

    usuario = Usuario.query.filter_by(email=email).first()
    if not usuario:
        flash('Erro: Usuário não encontrado no sistema.')
        return redirect(url_for('admin_login'))

    if request.method == 'POST':
        nova_senha = request.form.get('nova_senha')
        confirma_senha = request.form.get('confirma_senha')

        if nova_senha != confirma_senha:
            flash('Erro: As senhas digitadas não conferem.')
            return render_template('resetar_senha.html', token=token)

        # Salva a nova senha e encerra
        usuario.set_password(nova_senha)
        db.session.commit()
        flash('Sua senha foi redefinida com sucesso! Você já pode acessar o sistema.')
        return redirect(url_for('admin_login'))

    # Se for GET, mostra a telinha para digitar a nova senha
    return render_template('resetar_senha.html', token=token)

@app.route('/admin/dashboard', methods=['GET'])
def admin_dashboard():
    if 'user_id' not in session: return redirect(url_for('admin_login'))
    
    # Descobre quem está logado lendo a sessão
    user_login = session.get('user_login')
    user_sec_id = session.get('secretaria_id')
    
    # SE FOR O ADMIN: Vê todas as contratações e todas as secretarias no dropdown
    if user_login == 'admin':
        contratacoes = Contratacao.query.all()
        secretarias = Secretaria.query.all()
    # SE FOR USUÁRIO COMUM: Vê apenas da sua secretaria e só pode escolher a sua própria
    else:
        contratacoes = Contratacao.query.filter_by(secretaria_id=user_sec_id).all()
        secretarias = Secretaria.query.filter_by(id=user_sec_id).all()
        
    return render_template('admin_dashboard.html', contratacoes=contratacoes, secretarias=secretarias)

@app.route('/admin/cadastrar/contratacao', methods=['POST'])
def cadastrar_contratacao():
    if 'user_id' not in session: return redirect(url_for('admin_login'))
    
    user_login = session.get('user_login')
    user_sec_id = session.get('secretaria_id')
    form_sec_id = int(request.form.get('secretaria_id'))
    
    # VALIDAÇÃO DE SEGURANÇA BACKEND (RBAC)
    # Se não for o admin, e tentar mandar para outra secretaria, a gente bloqueia!
    if user_login != 'admin' and form_sec_id != user_sec_id:
        flash('Erro de Segurança: Você só pode cadastrar itens para a sua própria secretaria.')
        return redirect(url_for('admin_dashboard'))
        
    # === INÍCIO DA SANITIZAÇÃO DE DADOS (Lei de Postel) ===
    valor_raw = request.form.get('valor', '0')
        
    # 1. Remove 'R$', espaços em branco e deixa tudo limpo
    valor_clean = valor_raw.upper().replace('R$', '').strip()
    
    # 2. Se o usuário digitou no padrão BR com ponto de milhar e vírgula (ex: 1.000,00)
    if '.' in valor_clean and ',' in valor_clean:
        valor_clean = valor_clean.replace('.', '').replace(',', '.')
    # 3. Se digitou no padrão BR só com vírgula (ex: 1000,00)
    elif ',' in valor_clean:
        valor_clean = valor_clean.replace(',', '.')
    # 4. Se digitou padrão Americano (1000.00), deixamos como está, o float() aceita.
    
    try:
        valor_final = float(valor_clean)
    except ValueError:
        flash('Erro: O valor estimado inserido tem um formato inválido.')
        return redirect(url_for('admin_dashboard'))
    # === FIM DA SANITIZAÇÃO ===

    try:
        nova_contratacao = Contratacao(
            exercicio=int(request.form.get('exercicio')),
            objeto=request.form.get('objeto'),
            descricao=request.form.get('descricao'),
            valor_estimado=valor_final, # Passamos a variável sanitizada e convertida aqui!
            dotacao=request.form.get('dotacao'),
            data_planejada=datetime.strptime(request.form.get('data'), '%Y-%m-%d').date(),
            secretaria_id=int(request.form.get('secretaria_id'))
        )
        db.session.add(nova_contratacao)
        db.session.commit()
        flash('Item do PCA cadastrado com sucesso!') 
        
    except Exception as e:
        db.session.rollback()
        flash('Erro interno ao cadastrar a contratação. Verifique os dados e tente novamente.')
        print(f"Erro real do DB (Contratação): {str(e)}")
        
    return redirect(url_for('admin_dashboard'))

# ============================================================================
# EDITAR E EXCLUIR CONTRATAÇÕES
# ============================================================================

@app.route('/admin/editar/contratacao/<int:id>', methods=['POST'])
def editar_contratacao(id):
    if 'user_id' not in session: return redirect(url_for('admin_login'))
    
    contratacao = Contratacao.query.get_or_404(id)
    user_login = session.get('user_login')
    user_sec_id = session.get('secretaria_id')
    
    # SEGURANÇA RBAC: Verifica se é admin ou se o item pertence à secretaria do usuário
    if user_login != 'admin' and contratacao.secretaria_id != user_sec_id:
        flash('Erro: Acesso Negado. Você não pode alterar itens de outra secretaria.')
        return redirect(url_for('admin_dashboard'))

    # Sanitização do Valor (Lei de Postel - idêntica ao cadastro)
    valor_raw = request.form.get('valor', '0')
    valor_clean = valor_raw.upper().replace('R$', '').strip()
    if '.' in valor_clean and ',' in valor_clean: valor_clean = valor_clean.replace('.', '').replace(',', '.')
    elif ',' in valor_clean: valor_clean = valor_clean.replace(',', '.')

    try:
        valor_final = float(valor_clean)
        
        # Atualiza os dados no banco
        contratacao.exercicio = int(request.form.get('exercicio'))
        contratacao.objeto = request.form.get('objeto')
        contratacao.descricao = request.form.get('descricao')
        contratacao.valor_estimado = valor_final
        contratacao.dotacao = request.form.get('dotacao')
        contratacao.data_planejada = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
        
        # O Admin pode mudar a secretaria do item. O usuário comum não (mantém a dele).
        if user_login == 'admin' and request.form.get('secretaria_id'):
            contratacao.secretaria_id = int(request.form.get('secretaria_id'))

        db.session.commit()
        flash('Contratação atualizada com sucesso!')
    except Exception as e:
        db.session.rollback()
        flash('Erro interno ao atualizar a contratação.')
        print(f"Erro DB (Edição): {e}")

    return redirect(url_for('admin_dashboard'))

@app.route('/admin/excluir/contratacao/<int:id>', methods=['POST'])
def excluir_contratacao(id):
    if 'user_id' not in session: return redirect(url_for('admin_login'))
    
    contratacao = Contratacao.query.get_or_404(id)
    user_login = session.get('user_login')
    
    # SEGURANÇA RBAC
    if user_login != 'admin' and contratacao.secretaria_id != session.get('secretaria_id'):
        flash('Erro: Acesso Negado. Você não pode excluir itens de outra secretaria.')
        return redirect(url_for('admin_dashboard'))

    try:
        db.session.delete(contratacao)
        db.session.commit()
        flash('Contratação excluída com sucesso!')
    except Exception as e:
        db.session.rollback()
        flash('Erro interno ao excluir a contratação.')
        print(f"Erro DB (Exclusão): {e}")

    return redirect(url_for('admin_dashboard'))

# ============================================================================
# GESTÃO DE SECRETARIAS
# ============================================================================

@app.route('/admin/secretarias', methods=['GET'])
def gerenciar_secretarias():
    if 'user_id' not in session: return redirect(url_for('admin_login'))
    
    # TRAVA DE SEGURANÇA VERTICAL
    if session.get('user_login') != 'admin':
        flash('Acesso Negado: Apenas o administrador possui privilégios para gerenciar secretarias.')
        return redirect(url_for('admin_dashboard'))
    
    secretarias = Secretaria.query.all()
    return render_template('admin_secretarias.html', secretarias=secretarias)

@app.route('/admin/cadastrar/secretaria', methods=['POST'])
def cadastrar_secretaria():
    if 'user_id' not in session: return redirect(url_for('admin_login'))
    
    # TRAVA DE SEGURANÇA VERTICAL
    if session.get('user_login') != 'admin':
        flash('Acesso Negado: Você não tem permissão para executar esta ação.')
        return redirect(url_for('admin_dashboard'))

    nome_secretaria = request.form.get('nome')
    
    # Validação Proativa: Verifica se o nome da secretaria já existe no banco
    secretaria_existente = Secretaria.query.filter_by(nome=nome_secretaria).first()
    if secretaria_existente:
        flash(f'Erro: A secretaria "{nome_secretaria}" já está cadastrada no sistema.')
        return redirect(url_for('gerenciar_secretarias'))
        
    try:
        nova_sec = Secretaria(nome=nome_secretaria)
        db.session.add(nova_sec)
        db.session.commit()
        flash('Secretaria cadastrada com sucesso!')
        
    except Exception as e:
        db.session.rollback()
        # Oculta o erro real do banco de dados por segurança (Information Disclosure)
        flash('Erro interno ao cadastrar secretaria. Tente novamente mais tarde.')
        print(f"Erro real do DB (Secretaria): {str(e)}") # Fica apenas no log do servidor
        
    return redirect(url_for('gerenciar_secretarias'))


# ============================================================================
# GESTÃO DE USUÁRIOS
# ============================================================================

@app.route('/admin/usuarios', methods=['GET'])
def gerenciar_usuarios():
    if 'user_id' not in session: return redirect(url_for('admin_login'))
    
    # TRAVA DE SEGURANÇA VERTICAL
    if session.get('user_login') != 'admin':
        flash('Acesso Negado: Apenas o administrador possui privilégios para gerenciar usuários.')
        return redirect(url_for('admin_dashboard'))
    
    usuarios = Usuario.query.all()
    secretarias = Secretaria.query.all()
    return render_template('admin_usuarios.html', usuarios=usuarios, secretarias=secretarias)

@app.route('/admin/cadastrar/usuario', methods=['POST'])
def cadastrar_usuario():
    if 'user_id' not in session: return redirect(url_for('admin_login'))
    
    # TRAVA DE SEGURANÇA VERTICAL
    if session.get('user_login') != 'admin':
        flash('Acesso Negado: Você não tem permissão para executar esta ação.')
        return redirect(url_for('admin_dashboard'))

    login = request.form.get('login')
    senha = request.form.get('senha')
    confirma_senha = request.form.get('confirma_senha')
    email = request.form.get('email')

    # 1. Verifica se as senhas conferem
    if senha != confirma_senha:
        flash('Erro: As senhas digitadas não conferem. Tente novamente.')
        return redirect(url_for('gerenciar_usuarios')) 
        
    # 2. Verifica se o login já existe no banco (Prevenção do erro 1062)
    usuario_existente = Usuario.query.filter_by(login=login).first()
    if usuario_existente:
        flash(f'Erro: O login "{login}" já está em uso. Escolha um login diferente.')
        return redirect(url_for('gerenciar_usuarios'))

    # 3. Verifica se o e-mail já existe (Prevenção de erro UNIQUE) <--- NOVO
    if email:
        email_existente = Usuario.query.filter_by(email=email).first()
        if email_existente:
            flash(f'Erro: O e-mail "{email}" já está cadastrado para outro usuário.')
            return redirect(url_for('gerenciar_usuarios'))

    try:
        novo_user = Usuario(
            nome=request.form.get('nome'),
            login=login,
            email=email,
            secretaria_id=int(request.form.get('secretaria_id'))
        )
        novo_user.set_password(senha)
        
        db.session.add(novo_user)
        db.session.commit()
        flash('Usuário cadastrado com sucesso!')
        
    except Exception as e:
        db.session.rollback()
        # Mensagem genérica para o usuário, escondendo o erro do banco de dados
        flash('Erro interno ao cadastrar usuário. Tente novamente mais tarde.')
        print(f"Erro real do DB: {str(e)}") # O erro real fica apenas no terminal
        
    return redirect(url_for('gerenciar_usuarios'))

@app.route('/admin/alterar-senha', methods=['POST'])
def alterar_senha():
    if 'user_id' not in session: return redirect(url_for('admin_login'))
    
    usuario = Usuario.query.get(session['user_id'])
    senha_atual = request.form.get('senha_atual')
    nova_senha = request.form.get('nova_senha')
    confirma_senha = request.form.get('confirma_senha')
    
    # Verifica se o usuário sabe a senha atual antes de deixar mudar
    if not usuario.check_password(senha_atual):
        flash('Erro: A senha atual informada está incorreta.')
        return redirect(url_for('admin_dashboard'))
        
    if nova_senha != confirma_senha:
        flash('Erro: As novas senhas não conferem.')
        return redirect(url_for('admin_dashboard'))
        
    # Salva a nova senha encriptada e desloga o usuário por segurança
    usuario.set_password(nova_senha)
    db.session.commit()
    session.clear() 
    flash('Sua senha foi alterada com sucesso! Por favor, faça login novamente com a nova senha.')
    return redirect(url_for('admin_login'))

# ============================================================================
# GERENCIAMENTO DO ENTE (ADMIN)
# ============================================================================

@app.route('/admin/configuracoes', methods=['GET', 'POST'])
def configuracoes_ente():
    if session.get('user_login') != 'admin':
        flash('Erro: Apenas o Administrador pode alterar os dados do Órgão.')
        return redirect(url_for('admin_dashboard'))

    ente = Ente.query.first()
    if not ente:
        ente = Ente()
        db.session.add(ente)

    if request.method == 'POST':
        ente.nome = request.form.get('nome')
        ente.endereco = request.form.get('endereco')
        ente.telefone = request.form.get('telefone')
        ente.email = request.form.get('email')

        # Tratamento do Upload da Imagem
        logo = request.files.get('logo')
        if logo and logo.filename != '':
            filename = secure_filename(logo.filename)
            caminho_salvar = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            logo.save(caminho_salvar)
            ente.logo_path = f"img/{filename}" # Salva o caminho relativo no banco

        db.session.commit()
        flash('Configurações do Órgão atualizadas com sucesso!')
        return redirect(url_for('configuracoes_ente'))

    return render_template('admin_configuracoes.html', ente=ente)

# ============================================================================
# GERENCIAMENTO DE SECRETARIAS E USUÁRIOS (ADMIN)
# ============================================================================

@app.route('/admin/editar/secretaria/<int:id>', methods=['POST'])
def editar_secretaria(id):
    if session.get('user_login') != 'admin':
        flash('Erro: Acesso Negado.')
        return redirect(url_for('admin_dashboard'))

    sec = Secretaria.query.get_or_404(id)
    sec.nome = request.form.get('nome')
    db.session.commit()
    flash('Secretaria atualizada com sucesso!')
    return redirect(url_for('gerenciar_secretarias'))

@app.route('/admin/excluir/secretaria/<int:id>', methods=['POST'])
def excluir_secretaria(id):
    if session.get('user_login') != 'admin':
        flash('Erro: Acesso Negado.')
        return redirect(url_for('admin_dashboard'))

    sec = Secretaria.query.get_or_404(id)
    try:
        db.session.delete(sec)
        db.session.commit()
        flash('Secretaria excluída com sucesso!')
    except Exception as e:
        db.session.rollback()
        # Prevenção de erro de Chave Estrangeira (Integridade Referencial)
        flash('Erro ao excluir: Esta secretaria possui usuários ou contratações vinculadas. Exclua-os primeiro.')
    return redirect(url_for('gerenciar_secretarias'))

@app.route('/admin/editar/usuario/<int:id>', methods=['POST'])
def editar_usuario(id):
    if session.get('user_login') != 'admin':
        flash('Erro: Acesso Negado.')
        return redirect(url_for('admin_dashboard'))

    usuario = Usuario.query.get_or_404(id)
    novo_email = request.form.get('email') # <--- CAPTURANDO O E-MAIL

    # Validação: se o e-mail mudou, verificar se o novo já existe no banco <--- NOVO
    if novo_email and novo_email != usuario.email:
        email_existente = Usuario.query.filter_by(email=novo_email).first()
        if email_existente:
            flash(f'Erro: O e-mail "{novo_email}" já está sendo usado por outro usuário.')
            return redirect(url_for('gerenciar_usuarios'))

    try:
        usuario.nome = request.form.get('nome')
        usuario.login = request.form.get('login')
        usuario.email = novo_email # <--- ATUALIZANDO O DADO
        usuario.secretaria_id = int(request.form.get('secretaria_id'))
        
        db.session.commit()
        flash('Usuário atualizado com sucesso!')
    except Exception as e:
        db.session.rollback()
        flash('Erro interno ao atualizar usuário.')
        print(f"Erro real do DB (Edição de Usuário): {str(e)}")

    return redirect(url_for('gerenciar_usuarios'))

@app.route('/admin/excluir/usuario/<int:id>', methods=['POST'])
def excluir_usuario(id):
    if session.get('user_login') != 'admin':
        flash('Erro: Acesso Negado.')
        return redirect(url_for('admin_dashboard'))

    if id == session.get('user_id'):
        flash('Erro: Ação não permitida. Você não pode excluir a sua própria conta logada.')
        return redirect(url_for('gerenciar_usuarios'))

    usuario = Usuario.query.get_or_404(id)
    db.session.delete(usuario)
    db.session.commit()
    flash('Usuário excluído com sucesso!')
    return redirect(url_for('gerenciar_usuarios'))

@app.route('/admin/resetar-senha/usuario/<int:id>', methods=['POST'])
def resetar_senha_usuario(id):
    if session.get('user_login') != 'admin':
        flash('Erro: Acesso Negado.')
        return redirect(url_for('admin_dashboard'))

    usuario = Usuario.query.get_or_404(id)
    
    # 1. PROTEÇÃO CONTRA ESTAÇÃO DESBLOQUEADA (Zero Trust)
    # Se o admin tentar mudar a PRÓPRIA senha pela tabela, exige a senha atual
    if usuario.id == session.get('user_id'):
        senha_atual = request.form.get('senha_atual')
        if not usuario.check_password(senha_atual):
            flash('Erro de Segurança: A senha atual está incorreta. Operação bloqueada.')
            return redirect(url_for('gerenciar_usuarios'))

    nova_senha = request.form.get('nova_senha')
    confirma_senha = request.form.get('confirma_senha')
    
    # 2. VALIDAÇÃO DE REDUNDÂNCIA (Dupla digitação)
    if nova_senha != confirma_senha:
        flash('Erro: As novas senhas digitadas não conferem.')
        return redirect(url_for('gerenciar_usuarios'))
    
    usuario.set_password(nova_senha)
    db.session.commit()
    
    # 3. SE ELE MUDOU A PRÓPRIA SENHA, DESLOGA POR SEGURANÇA
    if usuario.id == session.get('user_id'):
        session.clear()
        flash('Sua senha foi alterada com sucesso! Por favor, faça login novamente com a nova credencial.')
        return redirect(url_for('admin_login'))

    flash(f'Senha do usuário {usuario.login} foi redefinida com sucesso!')
    return redirect(url_for('gerenciar_usuarios'))

if __name__ == '__main__':
    with app.app_context():
        # 1. Cria todas as tabelas no MySQL baseadas no models.py
        db.create_all()
        
        # 2. Verifica se existe pelo menos uma Secretaria (Obrigatório para o FK)
        sec_padrao = Secretaria.query.filter_by(nome='Secretaria de Administração').first()
        if not sec_padrao:
            sec_padrao = Secretaria(nome='Secretaria de Administração')
            db.session.add(sec_padrao)
            db.session.commit() # Commitamos aqui para gerar o ID da secretaria
            
        # 3. Verifica se o usuário 'admin' existe, se não, cria.
        admin_user = Usuario.query.filter_by(login='admin').first()
        if not admin_user:
            novo_admin = Usuario(
                nome='Administrador do Sistema',
                login='admin',
                secretaria_id=sec_padrao.id
            )

            # Puxa a senha do .env (Se não existir lá, usa uma senha aleatória gerada na hora para não travar)
            import secrets
            senha_inicial = os.environ.get('ADMIN_DEFAULT_PASSWORD', secrets.token_hex(8))
            novo_admin.set_password(senha_inicial)

            #novo_admin.set_password('admin123') # Usa a função de hash segura do models.py!
            
            db.session.add(novo_admin)
            db.session.commit()
            print("Bootstrapping concluído: Secretaria Padrão e Usuário Admin criados.")

    # Inicia o servidor Flask
    app.run(debug=True, host='127.0.0.1', port=5555)